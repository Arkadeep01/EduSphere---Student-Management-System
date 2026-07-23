import csv
import io
import os
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Callable, Optional

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.contrib.auth import get_user_model

from administration.models.export import ExportLog
from administration.models.admission import AdmissionApplication
from administration.models.contact import ContactSubmission
from administration.models.document import DocumentStorage
from administration.models.event import Event
from administration.models.exam import Exam, ExamSchedule, PublishedResult
from student.models import StudentProfile, StudentSubject, Attendance, Result
from teacher.models import TeacherProfile, TeacherClassAssignment, Resource, AnswerScript

User = get_user_model()


FIELD_MAP = {
    "name": {"label": "Name"},
    "username": {"label": "Username"},
    "email": {"label": "Email"},
    "phone": {"label": "Phone"},
    "gender": {"label": "Gender"},
    "date_of_birth": {"label": "Date of Birth"},
    "roll_number": {"label": "Roll Number"},
    "admission_number": {"label": "Admission Number"},
    "class_assigned": {"label": "Class"},
    "section": {"label": "Section"},
    "academic_session": {"label": "Academic Session"},
    "father_name": {"label": "Father Name"},
    "mother_name": {"label": "Mother Name"},
    "guardian_contact": {"label": "Guardian Contact"},
    "core_subjects": {"label": "Core Subjects"},
    "specialized_subjects": {"label": "Specialized Subjects"},
    "enriched_subjects": {"label": "Enriched Subjects"},
    "attendance_percentage": {"label": "Attendance Percentage"},
    "present_days": {"label": "Present Days"},
    "absent_days": {"label": "Absent Days"},
    "gpa": {"label": "GPA"},
    "overall_percentage": {"label": "Overall Percentage"},
    "rank": {"label": "Rank"},
    "assignment_average": {"label": "Assignment Average"},
    "exam_average": {"label": "Exam Average"},
    "fees_paid": {"label": "Fees Paid"},
    "fees_pending": {"label": "Fees Pending"},
    "fees_total_due": {"label": "Total Due"},
    "admission_status": {"label": "Admission Status"},
    "document_verification": {"label": "Document Verification"},
}


def _get_field_value(student, profile, field_key):
    user = student
    if field_key == "name":
        return user.get_full_name() or user.username
    if field_key == "username":
        return user.username
    if field_key == "email":
        return user.email
    if field_key == "phone":
        return getattr(user, "mobile", "")
    if field_key == "gender":
        return profile.gender or ""
    if field_key == "date_of_birth":
        return str(profile.date_of_birth or "")
    if field_key == "roll_number":
        return profile.roll_number or ""
    if field_key == "admission_number":
        return profile.admission_number or ""
    if field_key == "class_assigned":
        return profile.class_assigned or ""
    if field_key == "section":
        return profile.section or ""
    if field_key == "academic_session":
        return ""
    if field_key == "father_name":
        return profile.father_name or ""
    if field_key == "mother_name":
        return profile.mother_name or ""
    if field_key == "guardian_contact":
        return getattr(user, "mobile", "")
    if field_key == "core_subjects":
        subs = StudentSubject.objects.filter(student=profile, subject__tier="core", status="approved").select_related("subject")
        return ", ".join(s.subject.name for s in subs)
    if field_key == "specialized_subjects":
        subs = StudentSubject.objects.filter(student=profile, subject__tier="specialized", status="approved").select_related("subject")
        return ", ".join(s.subject.name for s in subs)
    if field_key == "enriched_subjects":
        subs = StudentSubject.objects.filter(student=profile, subject__tier="enrichment", status="approved").select_related("subject")
        return ", ".join(s.subject.name for s in subs)
    if field_key == "attendance_percentage":
        records = Attendance.objects.filter(student=profile)
        total = records.count()
        if total == 0:
            return "0%"
        present = records.filter(status="present").count()
        return f"{round(present / total * 100)}%"
    if field_key == "present_days":
        return str(Attendance.objects.filter(student=profile, status="present").count())
    if field_key == "absent_days":
        return str(Attendance.objects.filter(student=profile, status="absent").count())
    if field_key == "gpa":
        results = Result.objects.filter(student=profile)
        if not results:
            return ""
        total_ratio = sum(float(r.marks_obtained) / float(r.total_marks) for r in results)
        return f"{round((total_ratio / len(results)) * 4, 2)}"
    if field_key == "overall_percentage":
        results = Result.objects.filter(student=profile)
        if not results:
            return ""
        total_pct = sum(float(r.marks_obtained) / float(r.total_marks) * 100 for r in results)
        return f"{round(total_pct / len(results), 1)}%"
    if field_key == "rank":
        return ""
    if field_key == "assignment_average":
        from student.models import AssignmentSubmission
        subs = AssignmentSubmission.objects.filter(student=profile, grade__isnull=False)
        if not subs:
            return ""
        avg = sum(float(s.grade) for s in subs) / len(subs)
        return f"{round(avg, 1)}"
    if field_key == "exam_average":
        results = Result.objects.filter(student=profile)
        if not results:
            return ""
        avg = sum(float(r.marks_obtained) for r in results) / len(results)
        return f"{round(avg, 1)}"
    if field_key == "fees_paid":
        return ""
    if field_key == "fees_pending":
        return ""
    if field_key == "fees_total_due":
        return ""
    if field_key == "admission_status":
        return "Enrolled" if profile.class_assigned else "Pending"
    if field_key == "document_verification":
        docs = profile.admission_documents.all()
        return f"{docs.count()} document(s)"
    return ""


def _build_queryset(filters=None):
    qs = StudentProfile.objects.select_related("user").all()
    if filters:
        if filters.get("class_assigned"):
            qs = qs.filter(class_assigned=filters["class_assigned"])
        if filters.get("section"):
            qs = qs.filter(section=filters["section"])
        if filters.get("status"):
            qs = qs.filter(user__is_active=(filters["status"] == "active"))
        if filters.get("class_names"):
            qs = qs.filter(class_assigned__in=filters["class_names"])
        if filters.get("student_ids"):
            qs = qs.filter(id__in=filters["student_ids"])
    return qs


def _generate_filename(filters, file_format, prefix="Students"):
    parts = [prefix]
    if filters:
        if filters.get("class_assigned"):
            parts.append(filters["class_assigned"].replace(" ", ""))
        if filters.get("section"):
            parts.append(filters["section"])
        if filters.get("status"):
            parts.append(filters["status"].capitalize())
    parts.append(str(datetime.now().year))
    filename = "_".join(parts)
    ext_map = {"csv": "csv", "excel": "xlsx", "pdf": "pdf", "zip": "zip", "html": "html"}
    return f"{filename}.{ext_map.get(file_format, 'csv')}"


TEACHER_FIELD_MAP = {
    "name": {"label": "Name"},
    "employee_id": {"label": "Employee ID"},
    "username": {"label": "Username"},
    "email": {"label": "Email"},
    "phone": {"label": "Phone"},
    "gender": {"label": "Gender"},
    "qualification": {"label": "Qualification"},
    "experience": {"label": "Experience"},
    "date_of_joining": {"label": "Date Of Joining"},
    "employment_status": {"label": "Employment Status"},
    "assigned_subject": {"label": "Assigned Subject"},
    "assigned_classes": {"label": "Assigned Classes"},
    "assigned_sections": {"label": "Assigned Sections"},
    "class_teacher_of": {"label": "Class Teacher Of"},
    "total_students": {"label": "Total Students"},
    "pending_evaluations": {"label": "Pending Evaluations"},
    "assignments_created": {"label": "Assignments Created"},
    "attendance_taken": {"label": "Attendance Taken"},
    "resources_uploaded": {"label": "Resources Uploaded"},
    "syllabus_uploaded": {"label": "Syllabus Uploaded"},
    "last_login": {"label": "Last Login"},
    "performance_rating": {"label": "Performance Rating"},
}


def _get_teacher_field_value(user, profile, field_key):
    if field_key == "name":
        return user.get_full_name() or user.username
    if field_key == "employee_id":
        return profile.employee_id or ""
    if field_key == "username":
        return user.username
    if field_key == "email":
        return user.email
    if field_key == "phone":
        return getattr(user, "mobile", "")
    if field_key == "gender":
        return getattr(user, "gender", "")
    if field_key == "qualification":
        return profile.qualification or ""
    if field_key == "experience":
        return str(profile.experience) if profile.experience is not None else ""
    if field_key == "date_of_joining":
        return str(user.date_joined.date()) if user.date_joined else ""
    if field_key == "employment_status":
        return "Active" if user.is_active else "Inactive"
    if field_key == "assigned_subject":
        return profile.assigned_subject.name if profile.assigned_subject else ""
    if field_key == "assigned_classes":
        assignments = TeacherClassAssignment.objects.filter(teacher=profile)
        class_parts = sorted(set(a.class_name for a in assignments))
        return ", ".join(class_parts) if class_parts else ""
    if field_key == "assigned_sections":
        assignments = TeacherClassAssignment.objects.filter(teacher=profile)
        sections = sorted(set(a.class_name.split("-")[-1] if "-" in a.class_name else "" for a in assignments))
        sections = [s for s in sections if s]
        return ", ".join(sections) if sections else ""
    if field_key == "class_teacher_of":
        from administration.models.teacher import ClassTeacherAssignment
        cta = ClassTeacherAssignment.objects.filter(teacher=profile).first()
        return cta.class_name if cta else ""
    if field_key == "total_students":
        assignments = TeacherClassAssignment.objects.filter(teacher=profile)
        class_names = [a.class_name for a in assignments]
        if not class_names:
            return "0"
        total = StudentProfile.objects.filter(class_assigned__in=class_names).count()
        return str(total)
    if field_key == "pending_evaluations":
        return str(AnswerScript.objects.filter(teacher=profile, evaluation_status="pending").count())
    if field_key == "assignments_created":
        from student.models import Assignment
        return str(Assignment.objects.filter(created_by=user).count())
    if field_key == "attendance_taken":
        from student.models import Attendance
        return str(Attendance.objects.filter(marked_by=user).count())
    if field_key == "resources_uploaded":
        return str(Resource.objects.filter(teacher=profile).count())
    if field_key == "syllabus_uploaded":
        return ""
    if field_key == "last_login":
        return str(user.last_login) if user.last_login else ""
    if field_key == "performance_rating":
        return ""
    return ""


def _build_teacher_queryset(filters=None):
    qs = TeacherProfile.objects.select_related("user", "assigned_subject").all()
    if filters:
        if filters.get("subject_id"):
            qs = qs.filter(assigned_subject_id=filters["subject_id"])
        if filters.get("class_name"):
            qs = qs.filter(class_assignments__class_name=filters["class_name"])
        if filters.get("status") == "active":
            qs = qs.filter(user__is_active=True)
        elif filters.get("status") == "inactive":
            qs = qs.filter(user__is_active=False)
        if filters.get("class_teacher") is True:
            from administration.models.teacher import ClassTeacherAssignment
            teacher_ids = ClassTeacherAssignment.objects.values_list("teacher_id", flat=True)
            qs = qs.filter(id__in=teacher_ids)
        elif filters.get("class_teacher") is False:
            from administration.models.teacher import ClassTeacherAssignment
            teacher_ids = ClassTeacherAssignment.objects.values_list("teacher_id", flat=True)
            qs = qs.exclude(id__in=teacher_ids)
        if filters.get("teacher_ids"):
            qs = qs.filter(id__in=filters["teacher_ids"])
    return qs.distinct()


def _generate_teacher_filename(filters, file_format):
    parts = ["Teachers"]
    if filters:
        subject = filters.get("subject_name")
        if subject:
            parts.append(subject.replace(" ", ""))
        if filters.get("class_name"):
            parts.append(filters["class_name"].replace(" ", ""))
        if filters.get("class_teacher") is True:
            parts.append("ClassTeacher")
        elif filters.get("class_teacher") is False:
            parts.append("NonClassTeacher")
        if filters.get("status"):
            parts.append(filters["status"].capitalize())
    parts.append(str(datetime.now().year))
    filename = "_".join(parts)
    ext_map = {"csv": "csv", "excel": "xlsx", "pdf": "pdf"}
    return f"{filename}.{ext_map.get(file_format, 'csv')}"


# ---- CLASS FIELD MAP & HELPERS ----

CLASS_FIELD_MAP = {
    "name": {"label": "Class Name"},
    "section": {"label": "Section"},
    "total_students": {"label": "Total Students"},
    "class_teacher": {"label": "Class Teacher"},
    "subject_count": {"label": "Subject Count"},
    "academic_session": {"label": "Academic Session"},
}


def _get_class_field_value(class_name, section, filters, field_key):
    from administration.models.teacher import ClassTeacherAssignment
    if field_key == "name":
        return class_name
    if field_key == "section":
        return section or "-"
    if field_key == "total_students":
        return str(StudentProfile.objects.filter(class_assigned=class_name, section=section or "").count())
    if field_key == "class_teacher":
        cta = ClassTeacherAssignment.objects.filter(class_name=class_name).first()
        return cta.teacher.user.get_full_name() if cta else "-"
    if field_key == "subject_count":
        return "-"
    if field_key == "academic_session":
        return ""
    return ""


def _build_class_queryset(filters=None):
    from administration.models.teacher import ClassTeacherAssignment, TeacherSubjectAllocation
    classes_data = set()
    if filters:
        if filters.get("class_name"):
            classes_data.add((filters["class_name"], filters.get("section", "")))
        elif filters.get("class_names"):
            for cn in filters["class_names"]:
                classes_data.add((cn, ""))
        else:
            profiles = StudentProfile.objects.values("class_assigned", "section").distinct()
            for p in profiles:
                classes_data.add((p["class_assigned"], p["section"] or ""))
    else:
        profiles = StudentProfile.objects.values("class_assigned", "section").distinct()
        for p in profiles:
            classes_data.add((p["class_assigned"], p["section"] or ""))
    return sorted(classes_data)


# ---- EXAM FIELD MAP & HELPERS ----

EXAM_FIELD_MAP = {
    "name": {"label": "Exam Name"},
    "subject": {"label": "Subject"},
    "date": {"label": "Date"},
    "time": {"label": "Time"},
    "room": {"label": "Room"},
    "duration": {"label": "Duration"},
    "status": {"label": "Status"},
    "classes": {"label": "Classes"},
    "academic_year": {"label": "Academic Year"},
}


def _get_exam_field_value(exam, field_key):
    if field_key == "name":
        return exam.name
    if field_key == "subject":
        return exam.subject.name if exam.subject else ""
    if field_key == "date":
        return str(exam.date)
    if field_key == "time":
        return str(exam.time) if exam.time else ""
    if field_key == "room":
        return exam.room
    if field_key == "duration":
        return exam.duration
    if field_key == "status":
        return exam.status
    if field_key == "classes":
        return ", ".join(exam.classes or [])
    if field_key == "academic_year":
        return exam.academic_year
    return ""


def _build_exam_queryset(filters=None):
    qs = Exam.objects.select_related("subject").all()
    if filters:
        if filters.get("status"):
            qs = qs.filter(status=filters["status"])
        if filters.get("subject_id"):
            qs = qs.filter(subject_id=filters["subject_id"])
        if filters.get("academic_year"):
            qs = qs.filter(academic_year=filters["academic_year"])
    return qs


# ---- ADMISSION FIELD MAP & HELPERS ----

ADMISSION_FIELD_MAP = {
    "applicant_name": {"label": "Applicant Name"},
    "email": {"label": "Email"},
    "phone": {"label": "Phone"},
    "father_name": {"label": "Father Name"},
    "mother_name": {"label": "Mother Name"},
    "date_of_birth": {"label": "Date of Birth"},
    "address": {"label": "Address"},
    "previous_school": {"label": "Previous School"},
    "previous_board": {"label": "Previous Board"},
    "stream": {"label": "Stream"},
    "entrance_score": {"label": "Entrance Score"},
    "status": {"label": "Status"},
    "submitted_at": {"label": "Submitted At"},
}


def _get_admission_field_value(app, field_key):
    if field_key == "applicant_name":
        return app.applicant_name
    if field_key == "email":
        return app.applicant_email
    if field_key == "phone":
        return app.phone_number
    if field_key == "father_name":
        return app.father_name
    if field_key == "mother_name":
        return app.mother_name
    if field_key == "date_of_birth":
        return str(app.date_of_birth) if app.date_of_birth else ""
    if field_key == "address":
        return app.address
    if field_key == "previous_school":
        return app.previous_school
    if field_key == "previous_board":
        return app.previous_board
    if field_key == "stream":
        return app.stream
    if field_key == "entrance_score":
        if app.entrance_test_score is not None and app.entrance_test_total:
            return f"{app.entrance_test_score}/{app.entrance_test_total}"
        return ""
    if field_key == "status":
        return app.status
    if field_key == "submitted_at":
        return str(app.submitted_at.date()) if app.submitted_at else ""
    return ""


def _build_admission_queryset(filters=None):
    qs = AdmissionApplication.objects.all()
    if filters:
        if filters.get("status"):
            qs = qs.filter(status=filters["status"])
        if filters.get("stream"):
            qs = qs.filter(stream=filters["stream"])
    return qs


# ---- CONTACT FIELD MAP & HELPERS ----

CONTACT_FIELD_MAP = {
    "name": {"label": "Name"},
    "email": {"label": "Email"},
    "phone": {"label": "Phone"},
    "subject": {"label": "Subject"},
    "message": {"label": "Message"},
    "status": {"label": "Status"},
    "submitted_at": {"label": "Submitted At"},
}


def _get_contact_field_value(c, field_key):
    if field_key == "name":
        return c.name
    if field_key == "email":
        return c.email
    if field_key == "phone":
        return c.phone
    if field_key == "subject":
        return c.subject
    if field_key == "message":
        return c.message
    if field_key == "status":
        return c.status
    if field_key == "submitted_at":
        return str(c.submitted_at.date()) if c.submitted_at else ""
    return ""


def _build_contact_queryset(filters=None):
    qs = ContactSubmission.objects.all()
    if filters:
        if filters.get("status"):
            qs = qs.filter(status=filters["status"])
        if filters.get("archived") is not None:
            qs = qs.filter(archived=filters["archived"])
    return qs


# ---- AUDIT LOG FIELD MAP & HELPERS ----

AUDIT_FIELD_MAP = {
    "user": {"label": "User"},
    "action": {"label": "Action"},
    "model_name": {"label": "Model"},
    "object_id": {"label": "Object ID"},
    "description": {"label": "Description"},
    "ip_address": {"label": "IP Address"},
    "created_at": {"label": "Timestamp"},
}


def _get_audit_field_value(log, field_key):
    if field_key == "user":
        return log.user.email if log.user else ""
    if field_key == "action":
        return log.action
    if field_key == "model_name":
        return log.model_name
    if field_key == "object_id":
        return log.object_id
    if field_key == "description":
        return log.description
    if field_key == "ip_address":
        return str(log.ip_address) if log.ip_address else ""
    if field_key == "created_at":
        return str(log.created_at)
    return ""


def _build_audit_queryset(filters=None):
    from administration.models.audit_log import AuditLog
    qs = AuditLog.objects.select_related("user").all()
    if filters:
        if filters.get("action"):
            qs = qs.filter(action=filters["action"])
        if filters.get("model_name"):
            qs = qs.filter(model_name=filters["model_name"])
    return qs


# ---- DOCUMENT FIELD MAP & HELPERS ----

DOCUMENT_FIELD_MAP = {
    "original_filename": {"label": "File Name"},
    "file_type": {"label": "File Type"},
    "file_size": {"label": "File Size"},
    "uploaded_by": {"label": "Uploaded By"},
    "related_model": {"label": "Related Module"},
    "uploaded_at": {"label": "Uploaded At"},
}


def _get_document_field_value(doc, field_key):
    if field_key == "original_filename":
        return doc.original_filename
    if field_key == "file_type":
        return doc.file_type
    if field_key == "file_size":
        return f"{doc.file_size} bytes" if doc.file_size else ""
    if field_key == "uploaded_by":
        return doc.uploaded_by.email if doc.uploaded_by else ""
    if field_key == "related_model":
        return doc.related_model
    if field_key == "uploaded_at":
        return str(doc.uploaded_at.date()) if doc.uploaded_at else ""
    return ""


def _build_document_queryset(filters=None):
    qs = DocumentStorage.objects.select_related("uploaded_by").all()
    if filters:
        if filters.get("file_type"):
            qs = qs.filter(file_type=filters["file_type"])
        if filters.get("related_model"):
            qs = qs.filter(related_model=filters["related_model"])
    return qs


# ---- ATTENDANCE FIELD MAP & HELPERS ----

ATTENDANCE_FIELD_MAP = {
    "student": {"label": "Student"},
    "class_assigned": {"label": "Class"},
    "date": {"label": "Date"},
    "status": {"label": "Status"},
    "marked_by": {"label": "Marked By"},
}


def _get_attendance_field_value(a, field_key):
    if field_key == "student":
        return a.student.user.get_full_name() if a.student else ""
    if field_key == "class_assigned":
        return a.class_assigned or ""
    if field_key == "date":
        return str(a.date)
    if field_key == "status":
        return a.status
    if field_key == "marked_by":
        return a.marked_by.get_full_name() if a.marked_by else ""
    return ""


def _build_attendance_queryset(filters=None):
    qs = Attendance.objects.select_related("student__user", "marked_by").all()
    if filters:
        if filters.get("class_assigned"):
            qs = qs.filter(class_assigned=filters["class_assigned"])
        if filters.get("section"):
            qs = qs.filter(student__section=filters["section"])
        if filters.get("date_from"):
            qs = qs.filter(date__gte=filters["date_from"])
        if filters.get("date_to"):
            qs = qs.filter(date__lte=filters["date_to"])
        if filters.get("status"):
            qs = qs.filter(status=filters["status"])
    return qs


# ---- Export Config Registry ----

@dataclass
class ExportConfig:
    module_name: str
    label: str
    field_map: dict
    field_getter: Callable
    queryset_builder: Callable
    filename_generator: Callable = _generate_filename
    allowed_formats: list = field(default_factory=lambda: ["csv", "excel", "pdf"])


EXPORT_REGISTRY: dict[str, ExportConfig] = {}


def register_export(config: ExportConfig):
    EXPORT_REGISTRY[config.module_name] = config


register_export(ExportConfig(
    module_name="students",
    label="Students",
    field_map=FIELD_MAP,
    field_getter=lambda obj, f: _get_field_value(getattr(obj, "user", obj), obj, f),
    queryset_builder=_build_queryset,
    filename_generator=lambda flt, fmt: _generate_filename(flt, fmt, "Students"),
))

register_export(ExportConfig(
    module_name="teachers",
    label="Teachers",
    field_map=TEACHER_FIELD_MAP,
    field_getter=lambda obj, f: _get_teacher_field_value(getattr(obj, "user", obj), obj, f),
    queryset_builder=_build_teacher_queryset,
    filename_generator=_generate_teacher_filename,
))

register_export(ExportConfig(
    module_name="classes",
    label="Classes",
    field_map=CLASS_FIELD_MAP,
    field_getter=None,
    queryset_builder=_build_class_queryset,
    filename_generator=lambda flt, fmt: _generate_filename(flt, fmt, "Classes"),
))

register_export(ExportConfig(
    module_name="exams",
    label="Exams",
    field_map=EXAM_FIELD_MAP,
    field_getter=lambda obj, f: _get_exam_field_value(obj, f),
    queryset_builder=_build_exam_queryset,
    filename_generator=lambda flt, fmt: _generate_filename(flt, fmt, "Exams"),
))

register_export(ExportConfig(
    module_name="admissions",
    label="Admissions",
    field_map=ADMISSION_FIELD_MAP,
    field_getter=lambda obj, f: _get_admission_field_value(obj, f),
    queryset_builder=_build_admission_queryset,
    filename_generator=lambda flt, fmt: _generate_filename(flt, fmt, "Admissions"),
))

register_export(ExportConfig(
    module_name="contacts",
    label="Contact Forms",
    field_map=CONTACT_FIELD_MAP,
    field_getter=lambda obj, f: _get_contact_field_value(obj, f),
    queryset_builder=_build_contact_queryset,
    filename_generator=lambda flt, fmt: _generate_filename(flt, fmt, "Contacts"),
    allowed_formats=["pdf"],
))

register_export(ExportConfig(
    module_name="attendance",
    label="Attendance",
    field_map=ATTENDANCE_FIELD_MAP,
    field_getter=lambda obj, f: _get_attendance_field_value(obj, f),
    queryset_builder=_build_attendance_queryset,
    filename_generator=lambda flt, fmt: _generate_filename(flt, fmt, "Attendance"),
))

register_export(ExportConfig(
    module_name="audit_logs",
    label="Audit Logs",
    field_map=AUDIT_FIELD_MAP,
    field_getter=lambda obj, f: _get_audit_field_value(obj, f),
    queryset_builder=_build_audit_queryset,
    filename_generator=lambda flt, fmt: _generate_filename(flt, fmt, "AuditLogs"),
))

register_export(ExportConfig(
    module_name="documents",
    label="Documents",
    field_map=DOCUMENT_FIELD_MAP,
    field_getter=lambda obj, f: _get_document_field_value(obj, f),
    queryset_builder=_build_document_queryset,
    filename_generator=lambda flt, fmt: _generate_filename(flt, fmt, "Documents"),
))

# ---- FEE FIELD MAP & HELPERS ----

FEE_FIELD_MAP = {
    "student": {"label": "Student"},
    "class_assigned": {"label": "Class"},
    "month": {"label": "Month"},
    "total_fee": {"label": "Total Fee"},
    "paid_amount": {"label": "Paid Amount"},
    "fine": {"label": "Fine"},
    "gst": {"label": "GST"},
    "scholarship_amount": {"label": "Scholarship"},
    "status": {"label": "Status"},
    "payment_method": {"label": "Payment Method"},
    "receipt_number": {"label": "Receipt Number"},
    "paid_at": {"label": "Paid At"},
    "academic_session": {"label": "Academic Session"},
}


def _get_fee_field_value(payment, field_key):
    if field_key == "student":
        return payment.student.user.get_full_name() if payment.student else ""
    if field_key == "class_assigned":
        return payment.student.class_assigned or ""
    if field_key == "month":
        return payment.month
    if field_key == "total_fee":
        return str(payment.total_fee)
    if field_key == "paid_amount":
        return str(payment.paid_amount)
    if field_key == "fine":
        return str(payment.fine)
    if field_key == "gst":
        return str(payment.gst)
    if field_key == "scholarship_amount":
        return str(payment.scholarship_amount)
    if field_key == "status":
        return payment.status
    if field_key == "payment_method":
        return payment.payment_method or ""
    if field_key == "receipt_number":
        return payment.receipt_number or ""
    if field_key == "paid_at":
        return str(payment.paid_at.date()) if payment.paid_at else ""
    if field_key == "academic_session":
        return payment.academic_session
    return ""


def _build_fee_queryset(filters=None):
    from administration.models.fee import StudentFeePayment
    qs = StudentFeePayment.objects.select_related("student__user").all()
    if filters:
        if filters.get("class_assigned"):
            qs = qs.filter(student__class_assigned=filters["class_assigned"])
        if filters.get("month"):
            qs = qs.filter(month=filters["month"])
        if filters.get("status"):
            qs = qs.filter(status=filters["status"])
        if filters.get("academic_session"):
            qs = qs.filter(academic_session=filters["academic_session"])
    return qs


register_export(ExportConfig(
    module_name="fees",
    label="Fees",
    field_map=FEE_FIELD_MAP,
    field_getter=lambda obj, f: _get_fee_field_value(obj, f),
    queryset_builder=_build_fee_queryset,
    filename_generator=lambda flt, fmt: _generate_filename(flt, fmt, "Fees"),
))

# ---- RECEIPT FIELD MAP & HELPERS (filtered fee payments) ----


def _build_receipt_queryset(filters=None):
    from administration.models.fee import StudentFeePayment
    qs = StudentFeePayment.objects.select_related("student__user").filter(
        status="paid", receipt_number__isnull=False
    )
    if filters:
        if filters.get("class_assigned"):
            qs = qs.filter(student__class_assigned=filters["class_assigned"])
        if filters.get("month"):
            qs = qs.filter(month=filters["month"])
        if filters.get("receipt_number"):
            qs = qs.filter(receipt_number=filters["receipt_number"])
    return qs


register_export(ExportConfig(
    module_name="receipt",
    label="Receipts",
    field_map=FEE_FIELD_MAP,
    field_getter=lambda obj, f: _get_fee_field_value(obj, f),
    queryset_builder=_build_receipt_queryset,
    filename_generator=lambda flt, fmt: _generate_filename(flt, fmt, "Receipts"),
))


# ---- STAFF FIELD MAP & HELPERS ----

STAFF_FIELD_MAP = {
    "student_name": {"label": "Student Name"},
    "exam_name": {"label": "Exam Name"},
    "subject_name": {"label": "Subject Name"},
    "batch_id": {"label": "Batch ID"},
    "upload_status": {"label": "Upload Status"},
    "uploaded_by": {"label": "Uploaded By"},
    "uploaded_at": {"label": "Uploaded At"},
    "verified_by": {"label": "Verified By"},
    "verified_at": {"label": "Verified At"},
    "section": {"label": "Section"},
    "roll_number": {"label": "Roll Number"},
    "registration_number": {"label": "Registration Number"},
    "script_number": {"label": "Script Number"},
    "evaluation_status": {"label": "Evaluation Status"},
}


def _get_staff_field_value(script, field_key):
    if field_key == "student_name":
        return script.student.user.get_full_name() or script.student.user.email
    elif field_key == "exam_name":
        return script.exam.name
    elif field_key == "subject_name":
        return script.subject.name
    elif field_key == "batch_id":
        return script.batch_id
    elif field_key == "upload_status":
        return script.get_upload_status_display()
    elif field_key == "uploaded_by":
        return script.uploaded_by.email if script.uploaded_by else ""
    elif field_key == "uploaded_at":
        return str(script.uploaded_at.date()) if script.uploaded_at else ""
    elif field_key == "verified_by":
        return script.verified_by.email if script.verified_by else ""
    elif field_key == "verified_at":
        return str(script.verified_at.date()) if script.verified_at else ""
    elif field_key == "section":
        return script.section
    elif field_key == "roll_number":
        return script.roll_number
    elif field_key == "registration_number":
        return script.registration_number
    elif field_key == "script_number":
        return script.script_number
    elif field_key == "evaluation_status":
        return script.get_evaluation_status_display()
    return ""


def _build_staff_queryset(filters=None):
    from administration.models.exam import AnswerScriptUpload
    qs = AnswerScriptUpload.objects.select_related("exam", "subject", "student__user", "uploaded_by").all()
    if filters:
        if filters.get("upload_status"):
            qs = qs.filter(upload_status=filters["upload_status"])
        if filters.get("batch_id"):
            qs = qs.filter(batch_id=filters["batch_id"])
        if filters.get("exam_id"):
            qs = qs.filter(exam_id=filters["exam_id"])
    return qs


register_export(ExportConfig(
    module_name="staff",
    label="Staff Answer Scripts",
    field_map=STAFF_FIELD_MAP,
    field_getter=lambda obj, f: _get_staff_field_value(obj, f),
    queryset_builder=_build_staff_queryset,
    filename_generator=lambda flt, fmt: _generate_filename(flt, fmt, "StaffScripts"),
))

# ---- Core Export Service ----

class ExportService:

    # ---- Generic module export (config-driven) ----

    @staticmethod
    def export_module(user, module_name, file_format="csv", fields=None, filters=None):
        config = EXPORT_REGISTRY.get(module_name)
        if not config:
            raise ValueError(f"Unknown module: {module_name}")

        if file_format not in config.allowed_formats:
            raise ValueError(f"Format {file_format} not allowed for {module_name}")

        if config.module_name == "classes":
            return ExportService._export_classes(user, config, file_format, fields, filters)

        if fields is None:
            fields = list(config.field_map.keys())[:5]
        qs = config.queryset_builder(filters)

        if file_format == "excel":
            return ExportService._export_excel_generic(user, qs, fields, filters, config)
        if file_format == "pdf":
            return ExportService._export_pdf_generic(user, qs, fields, filters, config)
        return ExportService._export_csv_generic(user, qs, fields, filters, config)

    @staticmethod
    def _get_header(field_map, fields):
        return [field_map[f]["label"] for f in fields if f in field_map]

    @staticmethod
    def _get_rows(qs, fields, config):
        rows = []
        for obj in qs:
            row = []
            for f in fields:
                if f in config.field_map:
                    val = config.field_getter(obj, f)
                    row.append(str(val) if val is not None else "")
            rows.append(row)
        return rows

    @staticmethod
    def _export_csv_generic(user, qs, fields, filters, config):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(ExportService._get_header(config.field_map, fields))
        for row in ExportService._get_rows(qs, fields, config):
            writer.writerow(row)
        ExportLog.objects.create(
            user=user, model_name=config.module_name, export_type="csv",
            query_params={"fields": fields, "filters": filters or {}},
        )
        return output.getvalue()

    @staticmethod
    def _export_excel_generic(user, qs, fields, filters, config):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = config.label[:31]
        headers = ExportService._get_header(config.field_map, fields)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(ExportService._get_rows(qs, fields, config), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top")

        ws.auto_filter.ref = ws.dimensions
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        ExportLog.objects.create(
            user=user, model_name=config.module_name, export_type="excel",
            query_params={"fields": fields, "filters": filters or {}},
        )
        return output.getvalue()

    @staticmethod
    def _export_pdf_generic(user, qs, fields, filters, config):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output, pagesize=landscape(A4),
            topMargin=20 * mm, bottomMargin=20 * mm, leftMargin=15 * mm, rightMargin=15 * mm,
        )
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=18, spaceAfter=6, textColor=HexColor("#1e3a5f"))
        subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"], fontSize=10, spaceAfter=20, textColor=HexColor("#666666"))

        elements.append(Paragraph("EduSphere Academy", title_style))
        elements.append(Paragraph(
            f"{config.label} Export — Generated {datetime.now().strftime('%d %b %Y %H:%M')} by {user.get_full_name() or user.email}",
            subtitle_style,
        ))
        elements.append(Spacer(1, 6 * mm))

        headers = ExportService._get_header(config.field_map, fields)
        data_rows = [headers]
        data_rows.extend(ExportService._get_rows(qs, fields, config))

        if data_rows:
            col_width = (doc.width) / len(headers) if headers else 100
            table = Table(data_rows, colWidths=[col_width] * len(headers), repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#FFFFFF"), HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(table)

        elements.append(Spacer(1, 10 * mm))
        footer_style = ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=HexColor("#999999"))
        elements.append(Paragraph(
            f"EduSphere {config.label} Export — Page <page/> of <totalpages/> — Generated {datetime.now().strftime('%d-%b-%Y')}",
            footer_style,
        ))

        doc.build(elements)
        output.seek(0)

        ExportLog.objects.create(
            user=user, model_name=config.module_name, export_type="pdf",
            query_params={"fields": fields, "filters": filters or {}},
        )
        return output.getvalue()

    @staticmethod
    def _export_classes(user, config, file_format="csv", fields=None, filters=None):
        if fields is None:
            fields = ["name", "section", "total_students", "class_teacher", "subject_count"]
        classes_data = config.queryset_builder(filters)

        headers = ExportService._get_header(config.field_map, fields)
        data_rows = [headers]
        for (class_name, section) in classes_data:
            row = []
            for f in fields:
                if f in config.field_map:
                    row.append(str(_get_class_field_value(class_name, section, filters, f)))
            data_rows.append(row)

        if file_format == "pdf":
            return ExportService._export_pdf_from_rows(user, data_rows, config, fields, filters)
        if file_format == "excel":
            return ExportService._export_excel_from_rows(user, data_rows, config, fields, filters)
        return ExportService._export_csv_from_rows(user, data_rows, config, fields, filters)

    @staticmethod
    def _export_csv_from_rows(user, data_rows, config, fields, filters):
        output = io.StringIO()
        writer = csv.writer(output)
        for row in data_rows:
            writer.writerow(row)
        ExportLog.objects.create(
            user=user, model_name=config.module_name, export_type="csv",
            query_params={"fields": fields, "filters": filters or {}},
        )
        return output.getvalue()

    @staticmethod
    def _export_excel_from_rows(user, data_rows, config, fields, filters):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = config.label[:31]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

        for col_idx, header in enumerate(data_rows[0], 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(data_rows[1:], 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value).alignment = Alignment(vertical="top")

        ws.auto_filter.ref = ws.dimensions
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        ExportLog.objects.create(
            user=user, model_name=config.module_name, export_type="excel",
            query_params={"fields": fields, "filters": filters or {}},
        )
        return output.getvalue()

    @staticmethod
    def _export_pdf_from_rows(user, data_rows, config, fields, filters):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output, pagesize=landscape(A4),
            topMargin=20 * mm, bottomMargin=20 * mm, leftMargin=15 * mm, rightMargin=15 * mm,
        )
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=18, spaceAfter=6, textColor=HexColor("#1e3a5f"))
        subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"], fontSize=10, spaceAfter=20, textColor=HexColor("#666666"))

        elements.append(Paragraph("EduSphere Academy", title_style))
        elements.append(Paragraph(
            f"{config.label} Export — Generated {datetime.now().strftime('%d %b %Y %H:%M')} by {user.get_full_name() or user.email}",
            subtitle_style,
        ))
        elements.append(Spacer(1, 6 * mm))

        if data_rows:
            headers = data_rows[0]
            col_width = (doc.width) / len(headers) if headers else 100
            table = Table(data_rows, colWidths=[col_width] * len(headers), repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#FFFFFF"), HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(table)

        elements.append(Spacer(1, 10 * mm))
        footer_style = ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=HexColor("#999999"))
        elements.append(Paragraph(
            f"EduSphere {config.label} Export — Page <page/> of <totalpages/> — Generated {datetime.now().strftime('%d-%b-%Y')}",
            footer_style,
        ))

        doc.build(elements)
        output.seek(0)

        ExportLog.objects.create(
            user=user, model_name=config.module_name, export_type="pdf",
            query_params={"fields": fields, "filters": filters or {}},
        )
        return output.getvalue()

    # ---- Backward-compatible methods (students, teachers, attendance) ----

    @staticmethod
    def export_students(user, file_format="csv", fields=None, filters=None):
        if fields is None:
            fields = ["name", "email", "roll_number", "class_assigned", "section"]
        qs = _build_queryset(filters)
        config = EXPORT_REGISTRY["students"]
        if file_format == "excel":
            return ExportService._export_excel_generic(user, qs, fields, filters, config)
        if file_format == "pdf":
            return ExportService._export_pdf_generic(user, qs, fields, filters, config)
        return ExportService._export_csv_generic(user, qs, fields, filters, config)

    @staticmethod
    def export_teachers(user, file_format="csv", fields=None, filters=None):
        if fields is None:
            fields = ["name", "employee_id", "email", "assigned_subject", "experience"]
        qs = _build_teacher_queryset(filters)
        config = EXPORT_REGISTRY["teachers"]
        if file_format == "excel":
            return ExportService._export_excel_generic(user, qs, fields, filters, config)
        if file_format == "pdf":
            return ExportService._export_pdf_generic(user, qs, fields, filters, config)
        return ExportService._export_csv_generic(user, qs, fields, filters, config)

    @staticmethod
    def export_attendance(user, file_format="csv", fields=None, filters=None):
        if fields is None:
            fields = ["student", "class_assigned", "date", "status"]
        qs = _build_attendance_queryset(filters)
        config = EXPORT_REGISTRY["attendance"]
        if file_format == "pdf":
            return ExportService._export_pdf_generic(user, qs, fields, filters, config)
        if file_format == "excel":
            return ExportService._export_excel_generic(user, qs, fields, filters, config)
        return ExportService._export_csv_generic(user, qs, fields, filters, config)


# ---- ZIP Service ----

class ZIPService:

    @staticmethod
    def create_document_zip(queryset, filename_field="original_filename"):
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for obj in queryset:
                if hasattr(obj, "file") and obj.file:
                    file_path = obj.file.path if hasattr(obj.file, "path") else None
                    if file_path and os.path.exists(file_path):
                        arcname = getattr(obj, filename_field, None) or os.path.basename(file_path)
                        zf.write(file_path, arcname)
        buffer.seek(0)
        return buffer

    @staticmethod
    def create_admission_zip(admission_ids):
        from student.models import StudentProfile
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for aid in admission_ids:
                app = AdmissionApplication.objects.filter(id=aid).first()
                if not app:
                    continue
                profile = StudentProfile.objects.filter(admission_number=str(app.id)).first()
                if profile:
                    docs = profile.admission_documents.all()
                    for doc in docs:
                        if doc.document and os.path.exists(doc.document.path):
                            zf.write(doc.document.path, f"{app.applicant_name}/{os.path.basename(doc.document.path)}")
        buffer.seek(0)
        return buffer


# ---- Print Service ----

class PrintService:

    @staticmethod
    def generate_print_html(user, module_name, fields, filters):
        config = EXPORT_REGISTRY.get(module_name)
        if not config:
            raise ValueError(f"Unknown module: {module_name}")

        if config.module_name == "classes":
            classes_data = config.queryset_builder(filters)
            headers = ExportService._get_header(config.field_map, fields)
            rows = []
            for (class_name, section) in classes_data:
                row = []
                for f in fields:
                    if f in config.field_map:
                        row.append(str(_get_class_field_value(class_name, section, filters, f)))
                rows.append(row)
            data_rows = [headers] + rows
        else:
            if fields is None:
                fields = list(config.field_map.keys())[:5]
            qs = config.queryset_builder(filters)
            headers = ExportService._get_header(config.field_map, fields)
            rows = ExportService._get_rows(qs, fields, config)
            data_rows = [headers] + rows

        html = render_to_string("admin/export/print_layout.html", {
            "title": f"{config.label} Export",
            "generated_by": user.get_full_name() or user.email,
            "generated_at": datetime.now().strftime("%d %b %Y %H:%M"),
            "headers": headers,
            "rows": rows,
            "module_name": module_name,
        })
        return html
